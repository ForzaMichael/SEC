#!/usr/bin/env python3
"""Batch verification script for multiple Excel files."""

import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple
import json

import pandas as pd
from openpyxl import load_workbook

from sec_parser.client import SECClient
from sec_parser.extractor import FinancialExtractor, extract_metadata


class BatchVerifier:
    """Batch verification for multiple Excel files."""

    def __init__(self):
        self.client = SECClient()
        self.results = []

    def parse_excel_sheet(self, df: pd.DataFrame) -> Dict[str, float]:
        """Parse financial data from Excel sheet."""
        values = {}
        for idx, row in df.iterrows():
            if pd.notna(row[0]) and pd.notna(row[1]):
                label = str(row[0]).strip()
                try:
                    value = float(row[1])
                    values[label] = value
                except (ValueError, TypeError):
                    continue
        return values

    def read_excel_metadata(self, excel_path: str) -> Dict:
        """Extract metadata from Excel file."""
        wb = load_workbook(excel_path)
        metadata = {
            'file': excel_path,
            'sheets': wb.sheetnames,
            'company': None,
            'period_end': None,
            'fiscal_year': None,
        }

        # Try to extract from Income Statement
        if "Income Statement" in wb.sheetnames:
            ws = wb["Income Statement"]
            # Row 2 usually has company name
            if ws['A2'].value:
                metadata['company'] = str(ws['A2'].value)
            # Row 3 usually has period
            if ws['A3'].value:
                period_str = str(ws['A3'].value)
                metadata['period_end'] = period_str

        wb.close()
        return metadata

    def verify_file(self, excel_path: str, ticker: str, year: int) -> Dict:
        """Verify a single Excel file."""
        print(f"\n{'='*80}")
        print(f"Verifying: {excel_path}")
        print(f"{'='*80}")

        result = {
            'file': excel_path,
            'ticker': ticker,
            'year': year,
            'status': 'unknown',
            'total_items': 0,
            'matched_items': 0,
            'discrepancies': 0,
            'error': None,
        }

        try:
            # Read Excel data
            wb = load_workbook(excel_path)
            excel_data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(row)
                if rows:
                    excel_data[sheet_name] = pd.DataFrame(rows)
            wb.close()

            # Parse sheets
            excel_income = self.parse_excel_sheet(excel_data.get('Income Statement', pd.DataFrame()))
            excel_balance = self.parse_excel_sheet(excel_data.get('Balance Sheet', pd.DataFrame()))
            excel_cashflow = self.parse_excel_sheet(excel_data.get('Cash Flow', pd.DataFrame()))

            # Fetch SEC data
            cik = self.client.get_cik_for_ticker(ticker)
            if not cik:
                raise ValueError(f"Could not find CIK for {ticker}")

            filings = self.client.get_filings(cik, form_type="10-K")
            filing_info = None
            for f in filings:
                report_date = f.get('report_date', '')
                if report_date:
                    filing_year = int(report_date[:4])
                    if filing_year == year:
                        filing_info = f
                        break

            if not filing_info:
                raise ValueError(f"No 10-K filing found for year {year}")

            xbrl_facts = self.client.get_xbrl_facts(cik)
            metadata = extract_metadata(filing_info, xbrl_facts)

            from datetime import date, timedelta
            period_end = metadata.period_end
            period_start = date(period_end.year - 1, period_end.month, period_end.day) + timedelta(days=1)
            fiscal_year = period_end.year
            fiscal_period = "FY"

            extractor = FinancialExtractor(xbrl_facts, filing_info)
            sec_income = extractor.extract_income_statement(period_end, period_start, fiscal_year, fiscal_period, "10-K")
            sec_balance = extractor.extract_balance_sheet(period_end, fiscal_year, fiscal_period, "10-K")
            sec_cashflow = extractor.extract_cash_flow(period_end, period_start, fiscal_year, fiscal_period, "10-K")

            # Compare
            total_matches = 0
            total_items = 0
            discrepancies = 0

            for excel_vals, sec_stmt in [
                (excel_income, sec_income),
                (excel_balance, sec_balance),
                (excel_cashflow, sec_cashflow)
            ]:
                sec_values = {}
                for key, item in sec_stmt.line_items.items():
                    if item.value is not None:
                        if 'eps' in key.lower():
                            scaled_value = item.display_value
                        elif 'shares' in key.lower():
                            scaled_value = item.display_value / 1000
                        else:
                            scaled_value = item.display_value / 1000
                        sec_values[item.label] = scaled_value

                for label, excel_value in excel_vals.items():
                    if label in ['Line Item', 'ASSETS', 'LIABILITIES', "STOCKHOLDERS' EQUITY",
                                 'OPERATING ACTIVITIES', 'INVESTING ACTIVITIES', 'FINANCING ACTIVITIES']:
                        continue

                    if label in sec_values:
                        total_items += 1
                        sec_value = sec_values[label]
                        tolerance = abs(excel_value) * 0.001 + 0.01
                        diff = abs(excel_value - sec_value)

                        if diff > tolerance:
                            discrepancies += 1
                            print(f"❌ {label}: Excel={excel_value:,.2f}, SEC={sec_value:,.2f}, Diff={diff:,.2f}")
                        else:
                            total_matches += 1

            result['total_items'] = total_items
            result['matched_items'] = total_matches
            result['discrepancies'] = discrepancies
            result['status'] = 'passed' if discrepancies == 0 else 'failed'

            if discrepancies == 0:
                print(f"\n✅ PASSED: All {total_items} items matched")
            else:
                print(f"\n❌ FAILED: {discrepancies} discrepancies found out of {total_items} items")

        except Exception as e:
            result['status'] = 'error'
            result['error'] = str(e)
            print(f"\n❌ ERROR: {e}")

        return result

    def generate_summary_report(self, output_path: str = "batch_verification_summary.md"):
        """Generate a summary report of all verifications."""
        with open(output_path, 'w') as f:
            f.write("# 批量验证汇总报告\n\n")
            f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"**验证文件数**: {len(self.results)}\n\n")

            # Summary table
            f.write("## 验证结果汇总\n\n")
            f.write("| 文件 | 股票代码 | 财年 | 状态 | 匹配项 | 总项数 | 差异数 |\n")
            f.write("|------|----------|------|------|--------|--------|--------|\n")

            passed = 0
            failed = 0
            errors = 0

            for result in self.results:
                status_icon = {
                    'passed': '✅',
                    'failed': '❌',
                    'error': '⚠️'
                }.get(result['status'], '❓')

                f.write(f"| {Path(result['file']).name} | {result['ticker']} | {result['year']} | "
                       f"{status_icon} {result['status'].upper()} | {result['matched_items']} | "
                       f"{result['total_items']} | {result['discrepancies']} |\n")

                if result['status'] == 'passed':
                    passed += 1
                elif result['status'] == 'failed':
                    failed += 1
                else:
                    errors += 1

            # Statistics
            f.write(f"\n## 统计信息\n\n")
            f.write(f"- ✅ 通过: {passed}\n")
            f.write(f"- ❌ 失败: {failed}\n")
            f.write(f"- ⚠️ 错误: {errors}\n")
            f.write(f"- 📊 总计: {len(self.results)}\n\n")

            # Details
            f.write("## 详细信息\n\n")
            for result in self.results:
                f.write(f"### {Path(result['file']).name}\n\n")
                f.write(f"- **股票代码**: {result['ticker']}\n")
                f.write(f"- **财年**: {result['year']}\n")
                f.write(f"- **状态**: {result['status'].upper()}\n")
                f.write(f"- **匹配项**: {result['matched_items']}/{result['total_items']}\n")

                if result['error']:
                    f.write(f"- **错误**: {result['error']}\n")

                if result['status'] == 'passed':
                    f.write(f"- ✅ 所有数据验证通过\n")
                elif result['status'] == 'failed':
                    f.write(f"- ❌ 发现 {result['discrepancies']} 处差异\n")

                f.write("\n")

            f.write("---\n\n")
            f.write("*报告由批量验证脚本自动生成*\n")

        print(f"\n📄 Summary report saved to: {output_path}")


def main():
    """Main function."""
    if len(sys.argv) < 2:
        print("Usage: python3 batch_verify.py <file1.xlsx> [file2.xlsx] ...")
        print("   or: python3 batch_verify.py --dir <directory>")
        print("\nExamples:")
        print("  python3 batch_verify.py AAPL_10-K_2025.xlsx V_10-K_2025.xlsx")
        print("  python3 batch_verify.py --dir ./reports")
        sys.exit(1)

    verifier = BatchVerifier()

    # Collect files to verify
    files_to_verify = []

    if sys.argv[1] == '--dir':
        # Verify all Excel files in directory
        if len(sys.argv) < 3:
            print("Error: --dir requires a directory path")
            sys.exit(1)

        directory = Path(sys.argv[2])
        if not directory.is_dir():
            print(f"Error: {directory} is not a directory")
            sys.exit(1)

        for excel_file in directory.glob("*_10-K_*.xlsx"):
            files_to_verify.append(str(excel_file))
    else:
        # Verify specified files
        files_to_verify = sys.argv[1:]

    if not files_to_verify:
        print("No Excel files found to verify")
        sys.exit(1)

    print(f"Found {len(files_to_verify)} file(s) to verify")

    # Verify each file
    for excel_path in files_to_verify:
        # Extract ticker and year from filename
        filename = Path(excel_path).stem
        parts = filename.split('_')

        if len(parts) >= 3:
            ticker = parts[0]
            try:
                year = int(parts[-1])
            except ValueError:
                year = datetime.now().year
        else:
            print(f"⚠️  Skipping {excel_path}: Cannot parse ticker and year from filename")
            continue

        result = verifier.verify_file(excel_path, ticker, year)
        verifier.results.append(result)

    # Generate summary report
    print("\n" + "="*80)
    print("BATCH VERIFICATION COMPLETE")
    print("="*80)

    verifier.generate_summary_report()

    # Exit with appropriate code
    failed_count = sum(1 for r in verifier.results if r['status'] != 'passed')
    sys.exit(0 if failed_count == 0 else 1)


if __name__ == "__main__":
    main()
