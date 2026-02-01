#!/usr/bin/env python3
"""Detailed analysis to understand the Excel file data."""

import sys
from pathlib import Path
from openpyxl import load_workbook
from sec_parser.client import SECClient


def analyze_excel_file(excel_path: str):
    """Analyze Excel file to understand what data it contains."""
    print(f"Analyzing Excel file: {excel_path}\n")

    wb = load_workbook(excel_path)

    # Check Income Statement for metadata
    if "Income Statement" in wb.sheetnames:
        ws = wb["Income Statement"]

        print("=" * 80)
        print("EXCEL FILE METADATA")
        print("=" * 80)

        # Read first few rows to get metadata
        for i in range(1, 10):
            row_data = []
            for cell in ws[i]:
                if cell.value:
                    row_data.append(str(cell.value))
            if row_data:
                print(f"Row {i}: {' | '.join(row_data)}")

        print("\n" + "=" * 80)
        print("KEY FINANCIAL METRICS FROM EXCEL")
        print("=" * 80)

        # Find key metrics
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0] and row[1]:
                label = str(row[0]).strip()
                if label in ['Revenue', 'Net Income', 'EPS (Diluted)']:
                    print(f"{label}: {row[1]}")

    wb.close()


def fetch_multiple_years_sec_data(ticker: str):
    """Fetch SEC data for multiple years to find a match."""
    print("\n" + "=" * 80)
    print(f"SEC DATA FOR {ticker} - MULTIPLE YEARS")
    print("=" * 80)

    client = SECClient()

    # Get CIK
    cik = client.get_cik_for_ticker(ticker)
    if not cik:
        print(f"Could not find CIK for {ticker}")
        return

    print(f"CIK: {cik}\n")

    # Get recent filings
    filings = client.get_filings(cik, form_type="10-K", count=5)

    print(f"Found {len(filings)} recent 10-K filings:\n")

    for i, filing in enumerate(filings, 1):
        print(f"{i}. Filing Date: {filing.get('filing_date')}")
        print(f"   Report Date: {filing.get('report_date')}")
        print(f"   Accession: {filing.get('accession_number')}")

        # Get XBRL facts and extract key metrics
        try:
            xbrl_facts = client.get_xbrl_facts(cik)

            # Look for revenue in the facts
            us_gaap = xbrl_facts.get('facts', {}).get('us-gaap', {})

            # Try to find revenue for this filing
            revenue_concept = us_gaap.get('Revenues') or us_gaap.get('RevenueFromContractWithCustomerExcludingAssessedTax')
            if revenue_concept:
                units = revenue_concept.get('units', {})
                for unit_type, values in units.items():
                    for val in values:
                        if val.get('form') == '10-K':
                            report_date = val.get('end', '')
                            if report_date and report_date[:10] == filing.get('report_date'):
                                revenue = val.get('val', 0) / 1_000_000  # Convert to millions
                                print(f"   Revenue: ${revenue:,.0f}M")
                                break

            # Look for net income
            net_income_concept = us_gaap.get('NetIncomeLoss')
            if net_income_concept:
                units = net_income_concept.get('units', {})
                for unit_type, values in units.items():
                    for val in values:
                        if val.get('form') == '10-K':
                            report_date = val.get('end', '')
                            if report_date and report_date[:10] == filing.get('report_date'):
                                net_income = val.get('val', 0) / 1_000_000
                                print(f"   Net Income: ${net_income:,.0f}M")
                                break
        except Exception as e:
            print(f"   Error fetching details: {e}")

        print()


def main():
    if len(sys.argv) < 2:
        print("Usage: python detailed_analysis.py <excel_file>")
        sys.exit(1)

    excel_path = sys.argv[1]

    # Extract ticker from filename
    filename = Path(excel_path).stem
    ticker = filename.split('_')[0]

    # Analyze Excel file
    analyze_excel_file(excel_path)

    # Fetch SEC data for comparison
    fetch_multiple_years_sec_data(ticker)

    print("\n" + "=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)
    print("\nPlease compare the Excel file data with the SEC data above")
    print("to determine which fiscal year the Excel file actually represents.")


if __name__ == "__main__":
    main()
