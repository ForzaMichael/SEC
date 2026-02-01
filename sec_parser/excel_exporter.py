"""Excel export functionality for financial statements."""

from datetime import date
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import FinancialStatement, ParsedFiling, Segment, SegmentData, StatementType


class ExcelExporter:
    """Exports parsed SEC filing data to formatted Excel files."""

    def __init__(self):
        # Style definitions
        self.header_font = Font(bold=True, size=14)
        self.subheader_font = Font(bold=True, size=11)
        self.bold_font = Font(bold=True)
        self.number_format = '#,##0'
        self.number_format_decimal = '#,##0.00'
        self.percent_format = '0.00%'

        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, color="FFFFFF")
        self.subtotal_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        self.thin_border = Border(
            bottom=Side(style='thin', color='000000')
        )

    def export(self, parsed_filing: ParsedFiling, output_path: str) -> str:
        """
        Export parsed filing to Excel file.

        Args:
            parsed_filing: ParsedFiling object with all financial data
            output_path: Path for output Excel file

        Returns:
            Path to the created Excel file
        """
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active

        # Create sheets for each statement
        if parsed_filing.income_statement:
            self._write_income_statement(wb, parsed_filing)

        if parsed_filing.balance_sheet:
            self._write_balance_sheet(wb, parsed_filing)

        if parsed_filing.cash_flow_statement:
            self._write_cash_flow(wb, parsed_filing)

        if parsed_filing.segment_data:
            self._write_segments(wb, parsed_filing)

        # Remove default sheet if we created others
        if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Save workbook
        wb.save(output_path)
        return output_path

    def _write_header(
        self,
        ws,
        title: str,
        company_name: str,
        period_end: date,
        fiscal_period: str,
    ) -> int:
        """Write sheet header and return the next row number."""
        # Title
        ws['A1'] = title
        ws['A1'].font = self.header_font
        ws.merge_cells('A1:C1')

        # Company name
        ws['A2'] = company_name
        ws['A2'].font = self.subheader_font

        # Period
        period_str = f"Period Ending: {period_end.strftime('%B %d, %Y')}"
        if fiscal_period:
            period_str = f"{fiscal_period} - {period_str}"
        ws['A3'] = period_str

        # Amounts note
        ws['A4'] = "(Amounts in thousands, except per share data)"
        ws['A4'].font = Font(italic=True, size=9)

        return 6  # Return next row for data

    def _write_income_statement(self, wb: Workbook, parsed_filing: ParsedFiling) -> None:
        """Write income statement to a worksheet."""
        ws = wb.create_sheet("Income Statement")
        stmt = parsed_filing.income_statement
        meta = parsed_filing.metadata

        row = self._write_header(
            ws,
            "Consolidated Statement of Operations",
            meta.company_name,
            stmt.period_end,
            stmt.fiscal_period,
        )

        # Column headers
        ws.cell(row=row, column=1, value="Line Item").font = self.bold_font
        ws.cell(row=row, column=2, value="Amount").font = self.bold_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1

        # Define line item order and groupings
        line_order = [
            ("revenue", False),
            ("cost_of_revenue", False),
            ("gross_profit", True),
            ("research_and_development", False),
            ("selling_general_admin", False),
            ("operating_expenses", False),
            ("operating_income", True),
            ("interest_income", False),
            ("interest_expense", False),
            ("other_income_expense", False),
            ("income_before_tax", True),
            ("income_tax_expense", False),
            ("net_income", True),
            (None, False),  # Spacer
            ("eps_basic", False),
            ("eps_diluted", False),
            ("shares_basic", False),
            ("shares_diluted", False),
        ]

        row = self._write_line_items(ws, stmt, line_order, row)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _write_balance_sheet(self, wb: Workbook, parsed_filing: ParsedFiling) -> None:
        """Write balance sheet to a worksheet."""
        ws = wb.create_sheet("Balance Sheet")
        stmt = parsed_filing.balance_sheet
        meta = parsed_filing.metadata

        row = self._write_header(
            ws,
            "Consolidated Balance Sheet",
            meta.company_name,
            stmt.period_end,
            stmt.fiscal_period,
        )

        # Column headers
        ws.cell(row=row, column=1, value="Line Item").font = self.bold_font
        ws.cell(row=row, column=2, value="Amount").font = self.bold_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1

        # Assets section
        ws.cell(row=row, column=1, value="ASSETS").font = self.bold_font
        row += 1

        assets_order = [
            ("cash_and_equivalents", False),
            ("short_term_investments", False),
            ("accounts_receivable", False),
            ("inventory", False),
            ("prepaid_expenses", False),
            ("total_current_assets", True),
            ("property_plant_equipment", False),
            ("goodwill", False),
            ("intangible_assets", False),
            ("long_term_investments", False),
            ("other_assets", False),
            ("total_assets", True),
        ]

        row = self._write_line_items(ws, stmt, assets_order, row)
        row += 1

        # Liabilities section
        ws.cell(row=row, column=1, value="LIABILITIES").font = self.bold_font
        row += 1

        liabilities_order = [
            ("accounts_payable", False),
            ("accrued_liabilities", False),
            ("deferred_revenue_current", False),
            ("short_term_debt", False),
            ("current_portion_long_term_debt", False),
            ("total_current_liabilities", True),
            ("long_term_debt", False),
            ("deferred_tax_liabilities", False),
            ("other_liabilities", False),
            ("total_liabilities", True),
        ]

        row = self._write_line_items(ws, stmt, liabilities_order, row)
        row += 1

        # Equity section
        ws.cell(row=row, column=1, value="STOCKHOLDERS' EQUITY").font = self.bold_font
        row += 1

        equity_order = [
            ("common_stock", False),
            ("additional_paid_in_capital", False),
            ("retained_earnings", False),
            ("accumulated_other_comprehensive_income", False),
            ("treasury_stock", False),
            ("total_stockholders_equity", True),
            ("total_liabilities_and_equity", True),
        ]

        row = self._write_line_items(ws, stmt, equity_order, row)

        self._adjust_column_widths(ws)

    def _write_cash_flow(self, wb: Workbook, parsed_filing: ParsedFiling) -> None:
        """Write cash flow statement to a worksheet."""
        ws = wb.create_sheet("Cash Flow")
        stmt = parsed_filing.cash_flow_statement
        meta = parsed_filing.metadata

        row = self._write_header(
            ws,
            "Consolidated Statement of Cash Flows",
            meta.company_name,
            stmt.period_end,
            stmt.fiscal_period,
        )

        # Column headers
        ws.cell(row=row, column=1, value="Line Item").font = self.bold_font
        ws.cell(row=row, column=2, value="Amount").font = self.bold_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1

        # Operating activities
        ws.cell(row=row, column=1, value="OPERATING ACTIVITIES").font = self.bold_font
        row += 1

        operating_order = [
            ("net_income_cf", False),
            ("depreciation_amortization", False),
            ("stock_based_compensation", False),
            ("deferred_income_taxes", False),
            ("change_in_receivables", False),
            ("change_in_inventory", False),
            ("change_in_payables", False),
            ("other_operating_activities", False),
            ("net_cash_from_operating", True),
        ]

        row = self._write_line_items(ws, stmt, operating_order, row)
        row += 1

        # Investing activities
        ws.cell(row=row, column=1, value="INVESTING ACTIVITIES").font = self.bold_font
        row += 1

        investing_order = [
            ("capital_expenditures", False),
            ("acquisitions", False),
            ("purchases_of_investments", False),
            ("sales_of_investments", False),
            ("net_cash_from_investing", True),
        ]

        row = self._write_line_items(ws, stmt, investing_order, row)
        row += 1

        # Financing activities
        ws.cell(row=row, column=1, value="FINANCING ACTIVITIES").font = self.bold_font
        row += 1

        financing_order = [
            ("debt_issuance", False),
            ("debt_repayment", False),
            ("share_repurchases", False),
            ("dividends_paid", False),
            ("stock_issuance", False),
            ("net_cash_from_financing", True),
        ]

        row = self._write_line_items(ws, stmt, financing_order, row)
        row += 1

        # Net change
        net_change_order = [
            ("effect_of_exchange_rate", False),
            ("net_change_in_cash", True),
        ]

        row = self._write_line_items(ws, stmt, net_change_order, row)

        self._adjust_column_widths(ws)

    def _write_segments(self, wb: Workbook, parsed_filing: ParsedFiling) -> None:
        """Write segment data to a worksheet."""
        ws = wb.create_sheet("Segments")
        seg_data = parsed_filing.segment_data
        meta = parsed_filing.metadata

        row = self._write_header(
            ws,
            "Segment Performance",
            meta.company_name,
            seg_data.period_end,
            seg_data.fiscal_period,
        )

        # Business segments
        if seg_data.business_segments:
            ws.cell(row=row, column=1, value="BUSINESS SEGMENTS").font = self.bold_font
            row += 1

            # Headers
            headers = ["Segment", "Revenue", "Operating Income", "Assets"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.bold_font
                cell.fill = self.header_fill
                cell.font = self.header_font_white
            row += 1

            # Data rows
            for segment in seg_data.business_segments:
                ws.cell(row=row, column=1, value=segment.name)
                if segment.revenue:
                    cell = ws.cell(row=row, column=2, value=segment.revenue / 1000)
                    cell.number_format = self.number_format
                if segment.operating_income:
                    cell = ws.cell(row=row, column=3, value=segment.operating_income / 1000)
                    cell.number_format = self.number_format
                if segment.assets:
                    cell = ws.cell(row=row, column=4, value=segment.assets / 1000)
                    cell.number_format = self.number_format
                row += 1

            row += 1

        # Geographic segments
        if seg_data.geographic_segments:
            ws.cell(row=row, column=1, value="GEOGRAPHIC SEGMENTS").font = self.bold_font
            row += 1

            # Headers
            headers = ["Region", "Revenue", "Operating Income", "Assets"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.bold_font
                cell.fill = self.header_fill
                cell.font = self.header_font_white
            row += 1

            # Data rows
            for segment in seg_data.geographic_segments:
                ws.cell(row=row, column=1, value=segment.name)
                if segment.revenue:
                    cell = ws.cell(row=row, column=2, value=segment.revenue / 1000)
                    cell.number_format = self.number_format
                if segment.operating_income:
                    cell = ws.cell(row=row, column=3, value=segment.operating_income / 1000)
                    cell.number_format = self.number_format
                if segment.assets:
                    cell = ws.cell(row=row, column=4, value=segment.assets / 1000)
                    cell.number_format = self.number_format
                row += 1

        self._adjust_column_widths(ws)

    def _write_line_items(
        self,
        ws,
        stmt: FinancialStatement,
        line_order: List[Tuple[Optional[str], bool]],
        start_row: int,
    ) -> int:
        """
        Write line items to worksheet.

        Args:
            ws: Worksheet
            stmt: Financial statement
            line_order: List of (key, is_subtotal) tuples
            start_row: Starting row number

        Returns:
            Next row number after writing
        """
        row = start_row

        for key, is_subtotal in line_order:
            if key is None:
                # Spacer row
                row += 1
                continue

            if key not in stmt.line_items:
                continue

            item = stmt.line_items[key]

            # Label
            label_cell = ws.cell(row=row, column=1, value=item.label)
            if is_subtotal:
                label_cell.font = self.bold_font
                label_cell.fill = self.subtotal_fill

            # Value
            if item.value is not None:
                # Scale to thousands for display
                display_value = item.display_value
                if display_value is not None:
                    # EPS values should not be scaled
                    if 'eps' in key.lower():
                        scaled_value = display_value
                        num_format = self.number_format_decimal
                    elif 'shares' in key.lower():
                        scaled_value = display_value / 1000
                        num_format = self.number_format
                    else:
                        scaled_value = display_value / 1000
                        num_format = self.number_format

                    value_cell = ws.cell(row=row, column=2, value=scaled_value)
                    value_cell.number_format = num_format
                    value_cell.alignment = Alignment(horizontal='right')

                    if is_subtotal:
                        value_cell.font = self.bold_font
                        value_cell.fill = self.subtotal_fill
                        value_cell.border = self.thin_border

            row += 1

        return row

    def _adjust_column_widths(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Add some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
