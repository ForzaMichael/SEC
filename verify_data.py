#!/usr/bin/env python3
"""Verify Excel file data against SEC official data."""

import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from sec_parser.client import SECClient
from sec_parser.extractor import FinancialExtractor, extract_metadata


def read_excel_data(excel_path: str) -> Dict[str, pd.DataFrame]:
    """Read all sheets from Excel file."""
    print(f"Reading Excel file: {excel_path}")

    wb = load_workbook(excel_path)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract data from sheet
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)

        # Convert to DataFrame
        if rows:
            data[sheet_name] = pd.DataFrame(rows)

    wb.close()
    return data


def parse_excel_sheet(df: pd.DataFrame) -> Dict[str, float]:
    """Parse financial data from Excel sheet."""
    values = {}

    # Skip header rows and find data
    for idx, row in df.iterrows():
        if pd.notna(row[0]) and pd.notna(row[1]):
            label = str(row[0]).strip()
            try:
                value = float(row[1])
                values[label] = value
            except (ValueError, TypeError):
                continue

    return values


def fetch_sec_data(ticker: str, year: int) -> Tuple[Dict, Dict, Dict]:
    """Fetch data from SEC for comparison."""
    print(f"\nFetching SEC data for {ticker} (year {year})...")

    client = SECClient()

    # Get CIK
    cik = client.get_cik_for_ticker(ticker)
    if not cik:
        raise ValueError(f"Could not find CIK for {ticker}")

    print(f"CIK: {cik}")

    # Get filings
    filings = client.get_filings(cik, form_type="10-K")
    if not filings:
        raise ValueError(f"No 10-K filings found for {ticker}")

    # Find filing for the specified year
    filing_info = None
    for f in filings:
        report_date = f.get('report_date', '')
        if report_date:
            filing_year = int(report_date[:4])
            if filing_year == year:
                filing_info = f
                break

    if not filing_info:
        raise ValueError(f"No 10-K filing found for year {year}")

    print(f"Filing date: {filing_info.get('filing_date')}")
    print(f"Report date: {filing_info.get('report_date')}")

    # Get XBRL facts
    xbrl_facts = client.get_xbrl_facts(cik)

    # Extract metadata
    metadata = extract_metadata(filing_info, xbrl_facts)

    # Determine period dates
    period_end = metadata.period_end
    period_start = date(period_end.year - 1, period_end.month, period_end.day) + timedelta(days=1)
    fiscal_year = period_end.year
    fiscal_period = "FY"

    # Extract financial statements
    extractor = FinancialExtractor(xbrl_facts, filing_info)

    income_statement = extractor.extract_income_statement(
        period_end, period_start, fiscal_year, fiscal_period, "10-K"
    )

    balance_sheet = extractor.extract_balance_sheet(
        period_end, fiscal_year, fiscal_period, "10-K"
    )

    cash_flow = extractor.extract_cash_flow(
        period_end, period_start, fiscal_year, fiscal_period, "10-K"
    )

    return income_statement, balance_sheet, cash_flow


def compare_statements(excel_values: Dict[str, float], sec_statement, statement_name: str) -> List[str]:
    """Compare Excel values with SEC statement data."""
    print(f"\n{'='*80}")
    print(f"Comparing {statement_name}")
    print(f"{'='*80}")

    discrepancies = []
    matches = 0

    # Build mapping from SEC data
    sec_values = {}
    for key, item in sec_statement.line_items.items():
        if item.value is not None:
            # Scale to thousands like in Excel
            if 'eps' in key.lower():
                scaled_value = item.display_value
            elif 'shares' in key.lower():
                scaled_value = item.display_value / 1000
            else:
                scaled_value = item.display_value / 1000

            sec_values[item.label] = scaled_value

    # Compare each Excel value
    for label, excel_value in excel_values.items():
        # Skip header rows and non-data rows
        if label in ['Line Item', 'ASSETS', 'LIABILITIES', "STOCKHOLDERS' EQUITY",
                     'OPERATING ACTIVITIES', 'INVESTING ACTIVITIES', 'FINANCING ACTIVITIES']:
            continue

        if label in sec_values:
            sec_value = sec_values[label]

            # Allow small tolerance for rounding differences
            tolerance = abs(excel_value) * 0.001 + 0.01  # 0.1% + 0.01
            diff = abs(excel_value - sec_value)

            if diff > tolerance:
                discrepancy = f"❌ {label}:\n   Excel: {excel_value:,.2f}\n   SEC:   {sec_value:,.2f}\n   Diff:  {diff:,.2f}"
                print(discrepancy)
                discrepancies.append(discrepancy)
            else:
                print(f"✓ {label}: {excel_value:,.2f}")
                matches += 1
        else:
            print(f"⚠️  {label}: Not found in SEC data (Excel value: {excel_value:,.2f})")

    # Check for items in SEC but not in Excel
    for label, sec_value in sec_values.items():
        if label not in excel_values:
            print(f"⚠️  {label}: In SEC data but not in Excel (SEC value: {sec_value:,.2f})")

    print(f"\nMatches: {matches}")
    print(f"Discrepancies: {len(discrepancies)}")

    return discrepancies


def main():
    """Main verification function."""
    if len(sys.argv) < 2:
        print("Usage: python verify_data.py <excel_file> [year]")
        print("Example: python verify_data.py AAPL_10-K_2025.xlsx 2024")
        sys.exit(1)

    excel_path = sys.argv[1]

    # Extract ticker and year from filename
    filename = Path(excel_path).stem
    parts = filename.split('_')
    ticker = parts[0]

    if len(sys.argv) >= 3:
        year = int(sys.argv[2])
    else:
        # Try to extract from filename
        try:
            year = int(parts[-1])
        except (ValueError, IndexError):
            year = 2024

    print(f"Verifying data for {ticker} (year {year})")
    print(f"Excel file: {excel_path}")

    # Read Excel data
    excel_data = read_excel_data(excel_path)

    print(f"\nFound {len(excel_data)} sheets:")
    for sheet_name in excel_data.keys():
        print(f"  - {sheet_name}")

    # Parse Excel sheets
    excel_income = parse_excel_sheet(excel_data.get('Income Statement', pd.DataFrame()))
    excel_balance = parse_excel_sheet(excel_data.get('Balance Sheet', pd.DataFrame()))
    excel_cashflow = parse_excel_sheet(excel_data.get('Cash Flow', pd.DataFrame()))

    # Fetch SEC data
    sec_income, sec_balance, sec_cashflow = fetch_sec_data(ticker, year)

    # Compare statements
    all_discrepancies = []

    if excel_income:
        discrepancies = compare_statements(excel_income, sec_income, "Income Statement")
        all_discrepancies.extend(discrepancies)

    if excel_balance:
        discrepancies = compare_statements(excel_balance, sec_balance, "Balance Sheet")
        all_discrepancies.extend(discrepancies)

    if excel_cashflow:
        discrepancies = compare_statements(excel_cashflow, sec_cashflow, "Cash Flow Statement")
        all_discrepancies.extend(discrepancies)

    # Summary
    print(f"\n{'='*80}")
    print("VERIFICATION SUMMARY")
    print(f"{'='*80}")

    if all_discrepancies:
        print(f"\n❌ Found {len(all_discrepancies)} discrepancies:")
        for disc in all_discrepancies:
            print(f"\n{disc}")
        print("\n⚠️  Data verification FAILED - discrepancies found")
        sys.exit(1)
    else:
        print("\n✅ All data verified successfully!")
        print("Excel file data matches SEC official data.")
        sys.exit(0)


if __name__ == "__main__":
    main()
